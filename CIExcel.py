from openpyxl import Workbook
#from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import Alignment
import string

class CIWorkbook(Workbook):
    my_worksheets = {}
    working_sheet = None
    has_sheet = False

    def __init__(self):
        self.my_worksheets = {}
        self.working_sheet = None
        self.has_sheet = False
        super().__init__()

    def set_working_sheet(self,key):
        self.working_sheet = self.my_worksheets[key.lower()]

    def set_or_create_worksheet(self,key,columns=[],wrap=False):
        try: self.set_working_sheet(key)
        except: self.set_working_sheet(self.new_sheet(key,columns=columns,wrap=wrap))
        return key

    def new_sheet(self,title,columns=[],wrap=False):
        def define_sheet(key,num=1):
            if key in self.my_worksheets:
                return define_sheet(key+str(num),num+1)
            else:
                return key
        key = define_sheet(title.lower())
        if self.has_sheet:
            self.my_worksheets[key] = self.working_sheet = CIWorksheet(self,title=title,columns=columns,wrap=wrap)
        else:
            self.my_worksheets[key] = self.working_sheet = CIWorksheet(self,title=title,columns=columns, first_sheet=True,wrap=wrap)
            self.has_sheet = True
        #self.worksheets[key].title = title
        return key
    def new_summary(self,title):
        def define_sheet(key,num=1):
            if key in self.my_worksheets:
                return define_sheet(key+str(num),num+1)
            else:
                return key
        key = define_sheet(title.lower())
        if self.has_sheet:
            self.my_worksheets[key] = self.working_sheet = CITitlesheet(self,title=title)
        else:
            self.my_worksheets[key] = self.working_sheet = CITitlesheet(self,title=title,first_sheet=True)
            self.has_sheet = True
        return key

    def save(self,filename):
        for key,value in self.my_worksheets.items():
            all_columns = tuple(self.my_worksheets[key].sheet.columns)
            all_ascii = string.ascii_uppercase
            for col,current_ascii in zip(all_columns,all_ascii[0:len(all_columns)-1]):
                max_width = 1
                for row in col:
                    if row.value:
                        if len(str(row.value)) > max_width: max_width = len(str(row.value))
                #max_col_widths.append(max_width)
                #print(current_ascii)
                if max_width < 10:
                    self.my_worksheets[key].sheet.column_dimensions[current_ascii].width = 10
                else:
                    self.my_worksheets[key].sheet.column_dimensions[current_ascii].width = max_width + 1
        super().save(filename)

class CIWorksheet(object):
    title = None
    my_columns = []
    current_row = 2
    sheet = None
    wrap = False

    def __init__(self,wb,title=None,columns=[],first_sheet=False,wrap=False):
        self.wb = wb
        self.title = title
        self.my_columns = columns
        self.wrap = wrap

        if self.wb:
            if first_sheet:
                self.sheet = self.wb.active
                self.sheet.title = self.title
            else:
                self.sheet = self.wb.create_sheet(title=self.title)

        for i,c in enumerate(self.my_columns):
            self.sheet.cell(row=1,column=i+1,value=c).style = 'Headline 2'

    def add_row(self,input_row,row=None):
        for i,c in enumerate(self.my_columns):
            if c.lower() in input_row:
                if row:
                    self.sheet.cell(row=row,column=i+1,value=input_row[c.lower()])
                    try:
                        if "\n" in input_row[c.lower()]:
                            self.sheet.cell(row=row,column=i+1).alignment = Alignment(wrap_text=True)
                    except: pass

                else:
                    self.sheet.cell(row=self.current_row,column=i+1,value=input_row[c.lower()])
        if row:
            self.current_row = row + 1
        else:
            self.current_row += 1

class CITitlesheet(object):
    title = None
    current_row = 2
    list_group_count = 0
    sheet = None

    def __init__(self,wb,title=None,first_sheet=False,wrap=False):
        self.wb = wb
        self.title = title
        if not self.title:
            self.title = "Summary"

        if self.wb:
            if first_sheet:
                self.sheet = self.wb.active
                self.sheet.title = self.title
            else:
                self.sheet = self.wb.create_sheet(title=self.title)

        self.sheet.cell(row=1,column=1,value=self.title).style = 'Headline 1'

    def add_data(self,title,*args,format_value=None):
        column = 2
        for data in args:
            self.sheet.cell(row=self.current_row,column=1,value=title)
            if format_value:
                if type(data) is str and "%" in data:
                    self.sheet.cell(row=self.current_row,column=column,value=float(data.strip("%"))).number_format = format_value
                else:
                    self.sheet.cell(row=self.current_row,column=column,value=data)
            else:
                self.sheet.cell(row=self.current_row,column=column,value=data)
            column += 1
        self.current_row += 1

    def blank_data(self):
        self.current_row += 1

    def add_grouping(self,title,*args):
        self.list_group_count += 1
        column = (self.list_group_count * 4) + 1
        row = 2
        self.sheet.cell(row=1,column=column,value=title).style='Headline 3'
        for data in args:
            for i,d in enumerate(data):
                if type(d) is str and "%" in d:
                    self.sheet.cell(row=row,column=column+i,value=float(d.strip("%")))
                    self.sheet.cell(row=row,column=column+i).number_format = "0.00%"
                else:
                    self.sheet.cell(row=row,column=column+i,value=d)
            row += 1
        self.list_group_count += 1
